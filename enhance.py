import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load Excel file
file_path = 'Input.xlsx'
output54_df = pd.read_excel(file_path, sheet_name='output54', engine='openpyxl')
output97_df = pd.read_excel(file_path, sheet_name='output97', engine='openpyxl')

# Clean column headers
output54_df.columns = output54_df.columns.str.strip()
output97_df.columns = output97_df.columns.str.strip()

# Match by KEY column
matched_rows = []
matched_keys = set()

for idx54, row54 in output54_df.iterrows():
    key = row54['KEY']
    matching = output97_df[output97_df['KEY'] == key]
    if not matching.empty:
        row97 = matching.iloc[0]
        idx97 = matching.index[0]

        row54_meta = row54.copy()
        row54_meta['Source'] = 'output54'
        row54_meta['RowNumber'] = idx54 + 2

        row97_meta = row97.copy()
        row97_meta['Source'] = 'output97'
        row97_meta['RowNumber'] = idx97 + 2

        matched_rows.append(row54_meta)
        matched_rows.append(row97_meta)

        matched_keys.add(key)

# Unmatched rows
unmatched_rows = []

# From output54
for idx54, row54 in output54_df.iterrows():
    if row54['KEY'] not in matched_keys:
        row_unmatched = row54.copy()
        row_unmatched['Source'] = 'output54'
        row_unmatched['RowNumber'] = idx54 + 2
        unmatched_rows.append(row_unmatched)

# From output97
for idx97, row97 in output97_df.iterrows():
    if row97['KEY'] not in matched_keys:
        row_unmatched = row97.copy()
        row_unmatched['Source'] = 'output97'
        row_unmatched['RowNumber'] = idx97 + 2
        unmatched_rows.append(row_unmatched)

# Create comparison dataframe
comparison_df = pd.DataFrame(matched_rows)
cols = ['Source', 'RowNumber'] + [col for col in comparison_df.columns if col not in ['Source', 'RowNumber']]
comparison_df = comparison_df[cols]

# Save matched report to Excel
matched_file = 'comparison_report_with_highlights.xlsx'
comparison_df.to_excel(matched_file, index=False)

# Apply highlights to mismatches
wb = load_workbook(matched_file)
ws = wb.active
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for i in range(2, ws.max_row, 2):
    for j in range(3, ws.max_column + 1):  # skip Source and RowNumber
        cell1 = ws.cell(row=i, column=j)
        cell2 = ws.cell(row=i + 1, column=j)
        if cell1.value != cell2.value:
            cell1.fill = yellow_fill
            cell2.fill = yellow_fill

wb.save(matched_file)

# Save unmatched rows
unmatched_df = pd.DataFrame(unmatched_rows)
unmatched_cols = ['Source', 'RowNumber'] + [col for col in unmatched_df.columns if col not in ['Source', 'RowNumber']]
unmatched_df = unmatched_df[unmatched_cols]
unmatched_df.to_excel("unmatched_report.xlsx", index=False)

print("✓ Matched report: 'comparison_report_with_highlights.xlsx'")
print("✓ Unmatched rows saved in: 'unmatched_report.xlsx'")
